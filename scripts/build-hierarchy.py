# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl"]
# ///
"""Convert the official Census NAICS structure + descriptions files into naics-hierarchy.json.

Not in the beacon submodule (checked, absent) -> fetched separately per §C.
Structure: https://www.census.gov/naics/2022NAICS/2022_NAICS_Structure.xlsx
Descriptions (§R2): https://www.census.gov/naics/2022NAICS/2022_NAICS_Descriptions.xlsx
"""

import argparse
import json
import re
import urllib.request
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_URL = "https://www.census.gov/naics/2022NAICS/2022_NAICS_Structure.xlsx"
DEFAULT_DESCRIPTIONS_URL = "https://www.census.gov/naics/2022NAICS/2022_NAICS_Descriptions.xlsx"
HEADER_MARKER = "2022 NAICS Code"

# Descriptions xlsx jams definition + illustrative examples + cross-references
# into one free-text cell per code (§R2). We keep definition + examples,
# drop cross-references (§C: not surfaced in UI).
_EXAMPLES_MARKER = "Illustrative Examples:"
_CROSS_REFS_MARKER = "Cross-References."
_STUB_PREFIX = "See industry description for"

# census.gov 403s urllib's default User-Agent.
_USER_AGENT = "Mozilla/5.0 (compatible; naics-code-resolver build script)"


def fetch(url: str, dest: Path) -> None:
    req = urllib.request.Request(url, headers={"User-Agent": _USER_AGENT})
    with urllib.request.urlopen(req) as resp, dest.open("wb") as out:
        out.write(resp.read())


# Trilateral-agreement marker "T" is appended directly to the title with no
# separator (e.g. "Wheat FarmingT"). Strip it only when preceded by a
# lowercase letter, so a title that genuinely ends in a capital T survives.
_TRAILING_T = re.compile(r"(?<=[a-z])T$")


def clean_title(title: str) -> str:
    title = title.strip()
    return _TRAILING_T.sub("", title)


def parse_rows(raw_rows) -> list[tuple[str, str]]:
    """raw_rows: iterable of openpyxl row tuples. Returns [(code, title), ...] in file order."""
    rows = iter(raw_rows)
    for row in rows:
        if row and row[1] == HEADER_MARKER:
            break
    parsed = []
    for row in rows:
        code, title = row[1], row[2]
        if code is None or title is None:
            continue
        parsed.append((str(code), clean_title(str(title))))
    return parsed


def split_description(desc: str) -> tuple[str, list[str]] | None:
    """desc: raw 'Description' cell. Returns (definition, examples) with
    cross-references dropped, or None for a stub row ("See industry
    description for XXXXXX.") that carries no real content of its own."""
    desc = desc.strip()
    if desc.startswith(_STUB_PREFIX):
        return None
    before_refs = desc.split(_CROSS_REFS_MARKER, 1)[0]
    definition, _, examples_block = before_refs.partition(_EXAMPLES_MARKER)
    examples = [line.strip() for line in examples_block.splitlines() if line.strip()]
    return definition.strip(), examples


def parse_descriptions(raw_rows) -> dict[str, tuple[str, list[str]]]:
    """raw_rows: iterable of (code, title, description) tuples, header included."""
    rows = iter(raw_rows)
    next(rows)  # header: Code, Title, Description
    descriptions: dict[str, tuple[str, list[str]]] = {}
    for row in rows:
        code, _title, desc = row
        if code is None or not desc:
            continue
        parsed = split_description(str(desc))
        if parsed is not None:
            descriptions[str(code)] = parsed
    return descriptions


def merge_descriptions(tree: dict, descriptions: dict[str, tuple[str, list[str]]]) -> None:
    """Mutates tree in place, attaching definition/examples where §R2 has them."""
    for node in tree.values():
        found = descriptions.get(node["code"])
        if found is not None:
            definition, examples = found
            node["definition"] = definition
            if examples:
                node["examples"] = examples
        merge_descriptions(node["children"], descriptions)


def build_tree(parsed: list[tuple[str, str]]) -> dict:
    root: dict = {}
    level_stack: dict[int, dict] = {}
    for code, title in parsed:
        level = 2 if "-" in code else len(code)
        node = {"code": code, "title": title, "children": {}}
        parent = level_stack.get(level - 1)
        (parent["children"] if parent else root)[code] = node
        level_stack[level] = node
    return root


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--file", help="local structure xlsx path; skips the network fetch")
    parser.add_argument("--url", default=DEFAULT_URL)
    parser.add_argument("--descriptions-file", help="local descriptions xlsx path; skips the network fetch")
    parser.add_argument("--descriptions-url", default=DEFAULT_DESCRIPTIONS_URL)
    parser.add_argument("--skip-descriptions", action="store_true", help="titles only, no definitions/examples")
    parser.add_argument(
        "--out",
        default=str(
            REPO_ROOT / "packages" / "naics-search" / "src" / "data" / "naics-hierarchy.json"
        ),
    )
    args = parser.parse_args()

    import openpyxl

    if args.file:
        xlsx_path = Path(args.file)
    else:
        xlsx_path = REPO_ROOT / "scripts" / ".naics_structure.xlsx"
        fetch(args.url, xlsx_path)

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    parsed = parse_rows(wb.active.iter_rows(values_only=True))
    tree = build_tree(parsed)

    descriptions: dict[str, tuple[str, list[str]]] = {}
    if not args.skip_descriptions:
        if args.descriptions_file:
            desc_xlsx_path = Path(args.descriptions_file)
        else:
            desc_xlsx_path = REPO_ROOT / "scripts" / ".naics_descriptions.xlsx"
            fetch(args.descriptions_url, desc_xlsx_path)
        desc_wb = openpyxl.load_workbook(desc_xlsx_path, read_only=True, data_only=True)
        descriptions = parse_descriptions(desc_wb.active.iter_rows(values_only=True))
        merge_descriptions(tree, descriptions)

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(tree, separators=(",", ":")))
    print(f"wrote {out_path} ({len(parsed)} codes, {len(descriptions)} with definitions)")


if __name__ == "__main__":
    main()
